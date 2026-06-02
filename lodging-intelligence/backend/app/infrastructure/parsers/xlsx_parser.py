from __future__ import annotations

from pathlib import Path
from typing import Any

from app.infrastructure.parsers.base import ParsedDocument, ParsedPage, ParsedTable, rows_to_text


class XlsxParser:
    file_type = "xlsx"

    def parse(self, path: str | Path) -> ParsedDocument:
        from openpyxl import load_workbook

        workbook = load_workbook(Path(path), data_only=True, read_only=True)
        pages: list[ParsedPage] = []
        tables: list[ParsedTable] = []

        for sheet in workbook.worksheets:
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
            values = [row for row in values if any(cell not in (None, "") for cell in row)]
            if not values:
                continue

            headers = [
                str(value).strip() if value not in (None, "") else f"column_{index + 1}"
                for index, value in enumerate(values[0])
            ]
            rows: list[dict[str, Any]] = []
            for row in values[1:]:
                rows.append(
                    {
                        headers[index] if index < len(headers) else f"column_{index + 1}": value
                        for index, value in enumerate(row)
                    }
                )

            tables.append(
                ParsedTable(
                    source_name=f"{sheet.title}_table_1",
                    page_number=None,
                    sheet_name=sheet.title,
                    rows=rows,
                )
            )
            pages.append(
                ParsedPage(
                    page_number=None,
                    sheet_name=sheet.title,
                    section_name=sheet.title,
                    text=rows_to_text(rows),
                )
            )

        return ParsedDocument(
            file_type=self.file_type,
            pages=pages,
            tables=tables,
            metadata={"sheet_names": workbook.sheetnames},
        )
